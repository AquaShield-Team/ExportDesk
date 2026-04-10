"""
AQUASHIELD · ExportDesk — Motor ETL de Auditoría Comex v2.1
============================================================
Integra: SAP | Bill of Lading (multi-archivo) | Legalización DUS
Autor:   Equipo Comex Aquachile

Cambios v2.1 (Auditoría Técnica):
  - C-01: FECHA_HOY calculada dentro de procesar_auditoria(), no al importar
  - C-02: Columna Descripción SAP buscada dinámicamente (no hardcodeada)
  - A-01: limpiar_pedido retorna pd.NA para llaves vacías (evita cruces incorrectos)
  - A-02: determinar_estatus busca RESPONSABLE y RESPONSABLE_FINAL
  - A-05: Búsqueda de columna estado DUS más amplia y con avisos claros
  - M-01: Búsqueda de columna folio SAP más amplia
  - M-04: _aplicar_estilos diferencia tipos de excepción (ImportError, PermissionError, etc.)
"""

import pandas as pd
import glob
import os
from datetime import datetime

# ──────────────────────────────────────────────────────────
# CONFIGURACIÓN
# ──────────────────────────────────────────────────────────

FOLDER_PLANILLAS    = 'PLANILLAS'
SAP_FILE            = os.path.join(FOLDER_PLANILLAS, 'export.XLSX')
ASIGNACION_FILE     = os.path.join(FOLDER_PLANILLAS, '3.Asignacion Clientes.xlsx')
PROCESADOS_FILE     = os.path.join(FOLDER_PLANILLAS, '4.Procesados Manualmente.xlsx')
ANALISTAS_SAP_FILE  = os.path.join(FOLDER_PLANILLAS, '5.Analista que facturo.XLSX')
DUS_FILE_PATTERN    = os.path.join(FOLDER_PLANILLAS, '7. Legalización DUS*.xlsx')
OUTPUT_FILE         = 'Auditoria_Comex_Final.xlsx'

# NOTA C-01: FECHA_HOY ya NO se define aquí como variable global.
# Se calcula dentro de procesar_auditoria() para que cada ejecución
# use la fecha real del momento en que se llama la función.


# ──────────────────────────────────────────────────────────
# HELPERS
# ──────────────────────────────────────────────────────────

def limpiar_pedido(val):
    """
    Normaliza N° Pedido: elimina decimales, ceros iniciales y espacios.

    FIX A-01: Retorna pd.NA para valores que producen una llave vacía (""),
    evitando que filas sin pedido válido se crucen entre sí durante los merges.
    Casos que ahora retornan pd.NA:
      - NaN / None
      - Strings vacíos o solo espacios
      - "0", "00", "000" (quedan vacíos después de lstrip('0'))
    """
    if pd.isna(val):
        return pd.NA
    result = str(val).split('.')[0].strip().lstrip('0')
    return result if result else pd.NA


def calcular_semaforo(row, fecha_hoy):
    """
    Semáforo SLA basado en días desde fecha factura hasta hoy.

    FIX C-01: recibe fecha_hoy como parámetro (no usa variable global),
    garantizando que cada auditoría use la fecha del momento de ejecución.
    """
    fecha = row.get('FECHA FACTURA')
    if pd.isna(fecha):
        return "⚪ N/A"
    if fecha > fecha_hoy:
        return "💀 Rojo Crítico"       # Fecha futura = error de digitación SAP
    dias = (fecha_hoy - fecha).days
    if dias <= 3:   return "🟢 Verde"
    if dias <= 7:   return "🟠 Naranja"
    return "🔴 Rojo"


def determinar_estatus(row):
    """
    Árbol de decisión para ESTATUS_FINAL.
    Prioridad descendente:
    1. Sin factura    → Pendiente Facturación
    2. Error fecha    → Error Fecha
    3. DUS Legalizado → DUS Legalizado (cierre total del ciclo)
    4. BL subido      → Procesado
    5. Terrestre      → Procesado Terrestre
    6. Manual         → Validado Manualmente
    7. Default        → Pendiente

    FIX A-02: busca 'RESPONSABLE_FINAL' además de 'RESPONSABLE', porque
    en el momento en que se ejecuta apply(), el rename aún no se ha aplicado.
    """
    if pd.isna(row.get('Folio Factura')) or pd.isna(row.get('FECHA FACTURA')):
        return "📦 Pendiente Facturación"

    if row.get('SEMÁFORO') == "💀 Rojo Crítico":
        return "⚠️ Error Fecha"

    if row.get('DUS_LEGALIZADO', False):
        return "✅ DUS Legalizado"

    if pd.notna(row.get('BL')):
        return "✅ Procesado"

    # A-02: buscar en ambos nombres posibles (antes y después del rename de columnas)
    responsable_val = row.get('RESPONSABLE') or row.get('RESPONSABLE_FINAL', '')
    if 'Terrestre' in str(responsable_val):
        return "🚚 Procesado Terrestre"

    if pd.notna(row.get('MOTIVO')):
        return "⚠️ Validado Manualmente"

    return "❌ Pendiente"


def _aplicar_estilos(filepath):
    """
    Aplica colores de semáforo a la columna ESTATUS_FINAL en el Excel generado.

    FIX M-04: diferencia tipos de excepción para que el usuario sepa exactamente
    qué falló (librería no instalada vs. archivo abierto vs. error inesperado).
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import PatternFill, Font, Alignment
    except ImportError:
        print("    Aviso: openpyxl no instalado. Instala con: pip install openpyxl")
        return

    try:
        COLORES = {
            'DUS Leg': ('D4EDDA', '155724'),
            '✅':      ('C6EFCE', '006100'),
            '🚚':      ('DEEBF7', '003366'),
            '⚠️':     ('FFF2CC', '9C6500'),
            '📦':      ('FFF2CC', '9C6500'),
            '❌':      ('FFC7CE', '9C0006'),
        }

        wb = load_workbook(filepath)
        ws = wb.active

        header = [cell.value for cell in ws[1]]
        try:
            status_idx = header.index('ESTATUS_FINAL') + 1
        except ValueError:
            status_idx = 1

        for row in range(2, ws.max_row + 1):
            cell = ws.cell(row=row, column=status_idx)
            val  = str(cell.value or '')
            bg, fg = 'FFFFFF', '000000'
            if 'DUS Legalizado' in val: bg, fg = COLORES['DUS Leg']
            elif '✅' in val:           bg, fg = COLORES['✅']
            elif '🚚' in val:           bg, fg = COLORES['🚚']
            elif '⚠️' in val:          bg, fg = COLORES['⚠️']
            elif '📦' in val:           bg, fg = COLORES['📦']
            elif '❌' in val:           bg, fg = COLORES['❌']
            cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
            cell.font = Font(color=fg, bold=True)
            cell.alignment = Alignment(vertical='center')

        for col in ws.columns:
            max_len = max((len(str(c.value or '')) for c in col), default=0)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

        wb.save(filepath)
        print("    Estilos aplicados correctamente.")

    except PermissionError:
        print(f"    ⚠️  El archivo '{filepath}' está abierto en Excel. Ciérralo e intenta de nuevo.")
    except Exception as e:
        print(f"    ⚠️  Estilizado omitido — {type(e).__name__}: {e}")


# ──────────────────────────────────────────────────────────
# MOTOR PRINCIPAL
# ──────────────────────────────────────────────────────────

def procesar_auditoria():
    # C-01: FECHA_HOY se calcula aquí, en el momento de ejecutar la auditoría
    FECHA_HOY = datetime.today().replace(hour=0, minute=0, second=0, microsecond=0)

    print(f"AQUASHIELD · ExportDesk — Auditoría {FECHA_HOY.strftime('%d/%m/%Y')}")
    print("=" * 60)

    # ── 1. SAP ────────────────────────────────────────────
    print("[1/6] Cargando SAP...")
    df_sap = pd.read_excel(SAP_FILE)
    col_map = {col: col.encode('ascii', 'ignore').decode('ascii') for col in df_sap.columns}
    df_sap = df_sap.rename(columns=col_map)

    # C-02: Buscar columna de descripción dinámicamente (no hardcodeada como 'Descripcin')
    desc_col = next(
        (c for c in df_sap.columns if 'descrip' in c.lower() or 'material' in c.lower()),
        None
    )
    if desc_col is None:
        raise ValueError(
            "No se encontró columna de Descripción en el reporte SAP.\n"
            f"    Columnas disponibles: {list(df_sap.columns)}"
        )
    print(f"    → Columna de descripción detectada: '{desc_col}'")

    # Filtros de seguridad: @ = Congelado (incluir), # = Fresco (excluir)
    df_sap = df_sap[df_sap[desc_col].str.contains('@', na=False)]
    df_sap = df_sap[~df_sap[desc_col].str.contains('#', na=False)]

    df_sap['PEDIDO_CLEAN'] = df_sap['NPedido'].apply(limpiar_pedido)
    # A-01: descartar filas con pedido vacío/inválido antes de deduplicar
    df_sap = df_sap.dropna(subset=['PEDIDO_CLEAN'])
    df_sap = df_sap.drop_duplicates(subset=['PEDIDO_CLEAN'], keep='first')
    print(f"    → {len(df_sap)} pedidos SAP (post-filtro)")

    # ── 2. BL (multi-archivo) ─────────────────────────────
    print("[2/6] Cargando Bill of Ladings...")
    bl_files = glob.glob(os.path.join(FOLDER_PLANILLAS, 'bills_of_lading-*.xlsx'))
    if bl_files:
        df_bl = pd.concat([pd.read_excel(f) for f in bl_files], ignore_index=True)
        df_bl['PEDIDO_CLEAN'] = df_bl['Pedido'].apply(limpiar_pedido)
        # A-01: descartar BLs sin pedido válido
        df_bl = df_bl.dropna(subset=['PEDIDO_CLEAN'])
        df_bl = df_bl.drop_duplicates(subset=['PEDIDO_CLEAN'], keep='last')  # más reciente
        print(f"    → {len(bl_files)} archivo(s) BL, {len(df_bl)} registros únicos")
    else:
        print("    → ADVERTENCIA: No se encontraron archivos BL.")
        df_bl = pd.DataFrame(columns=['PEDIDO_CLEAN', 'BL', 'FechaCreacion'])

    # ── 3. Maestros adicionales ───────────────────────────
    print("[3/6] Cargando maestros...")
    df_asignacion = pd.read_excel(ASIGNACION_FILE)
    df_manual     = pd.read_excel(PROCESADOS_FILE)
    df_manual['PEDIDO_CLEAN'] = df_manual['PEDIDO FLUJO'].apply(limpiar_pedido)
    df_manual = df_manual.dropna(subset=['PEDIDO_CLEAN'])  # A-01

    df_analistas = pd.DataFrame()
    if os.path.exists(ANALISTAS_SAP_FILE):
        df_analistas = pd.read_excel(ANALISTAS_SAP_FILE)
        # M-01: búsqueda más amplia de columna folio (agrega 'Doc.' y 'Documento')
        factura_cols = [c for c in df_analistas.columns
                        if any(kw in c for kw in ['Factura', 'Folio', 'Doc.', 'Documento'])]
        pedido_cols  = [c for c in df_analistas.columns
                        if any(kw in c for kw in ['Pedido', 'Orden', 'NPedido'])]
        if factura_cols:
            df_analistas['FOLIO_JOIN'] = df_analistas[factura_cols[0]].apply(limpiar_pedido)
        if pedido_cols:
            df_analistas['PEDIDO_JOIN'] = df_analistas[pedido_cols[0]].apply(limpiar_pedido)
        # P-07: limpiar NaN en AMBAS columnas de cruce (no solo FOLIO_JOIN)
        # Sin esto, si factura_cols está vacío pero pedido_cols existe,
        # PEDIDO_JOIN puede contener pd.NA y generar filas fantasma en el merge.
        dropna_cols = []
        if factura_cols: dropna_cols.append('FOLIO_JOIN')
        if pedido_cols:  dropna_cols.append('PEDIDO_JOIN')
        if dropna_cols:
            df_analistas = df_analistas.dropna(subset=dropna_cols, how='all')  # A-01 + P-07

    # ── 4. DUS Legalizado ─────────────────────────────────
    print("[4/6] Cargando reporte DUS...")
    dus_files = glob.glob(DUS_FILE_PATTERN)
    if dus_files:
        dus_files.sort(key=os.path.getmtime, reverse=True)  # más reciente
        df_dus = pd.read_excel(dus_files[0])
        df_dus['PEDIDO_CLEAN'] = df_dus['N° PEDIDO'].apply(limpiar_pedido)
        df_dus = df_dus.dropna(subset=['PEDIDO_CLEAN'])  # A-01

        # A-05: búsqueda de columna estado DUS más amplia, con avisos claros
        estado_col = next(
            (c for c in df_dus.columns if 'ESTADO' in c.upper() and 'GENERAL' in c.upper()),
            None
        )
        if estado_col is None:
            # Segundo intento: cualquier columna con ESTADO y DUS
            estado_col = next(
                (c for c in df_dus.columns if 'ESTADO' in c.upper() and 'DUS' in c.upper()),
                None
            )
        if estado_col is None:
            # Último intento: cualquier columna con ESTADO
            estado_col = next((c for c in df_dus.columns if 'ESTADO' in c.upper()), None)
            if estado_col:
                print(f"    ⚠️  AVISO: usando '{estado_col}' para estado DUS (nombre no estándar).")
            else:
                print(f"    ❌ ERROR: No se encontró columna de estado DUS.")
                print(f"    Columnas disponibles: {list(df_dus.columns)}")

        fecha_col   = next((c for c in df_dus.columns if 'FECHA' in c.upper()), None)
        detalle_col = next((c for c in df_dus.columns if 'DETALLE' in c.upper()), None)

        df_dus['DUS_LEGALIZADO'] = (
            df_dus[estado_col].str.contains('✅', na=False) if estado_col else False
        )
        df_dus['FECHA_DUS']      = (
            pd.to_datetime(df_dus[fecha_col], errors='coerce') if fecha_col else pd.NaT
        )
        df_dus['ESTADO_DUS_RAW'] = df_dus[detalle_col] if detalle_col else ''

        # Si mismo pedido aparece legalizado y pendiente, legalizado gana
        df_dus = df_dus.sort_values('DUS_LEGALIZADO', ascending=False)
        df_dus = df_dus.drop_duplicates(subset=['PEDIDO_CLEAN'], keep='first')
        print(f"    → {len(df_dus)} registros DUS ({df_dus['DUS_LEGALIZADO'].sum()} legalizados)")
    else:
        print("    → ADVERTENCIA: No se encontró archivo DUS.")
        df_dus = pd.DataFrame(columns=['PEDIDO_CLEAN', 'DUS_LEGALIZADO', 'FECHA_DUS', 'ESTADO_DUS_RAW'])
        df_dus['DUS_LEGALIZADO'] = pd.Series(dtype=bool)

    # ── 5. Cruces en cascada ──────────────────────────────
    print("[5/6] Realizando cruces...")
    df_sap['CLIENTE_NORM']        = df_sap['Nombre Solicitante'].astype(str).str.lower().str.strip()
    df_asignacion['CLIENTE_NORM'] = df_asignacion['Cliente'].astype(str).str.lower().str.strip()

    # P-03: deduplicar asignación por CLIENTE_NORM ANTES del merge
    # Si un cliente aparece 2 veces con responsables distintos, el merge
    # multiplicaría las filas del SAP (ej: 3000 pedidos × 2 = 6000 filas basura).
    df_asignacion = df_asignacion.drop_duplicates(subset=['CLIENTE_NORM'], keep='first')

    # P-02: set en vez de list → O(1) per lookup en calcular_responsable
    # Con 10,000 pedidos y 500 clientes terrestres, esto pasa de ~500ms a <1ms.
    terrestres_set = set(
        df_asignacion[
            df_asignacion['Responsable'].str.contains('Terrestre', na=False, case=False)
        ]['CLIENTE_NORM']
    )

    # M-01: búsqueda más amplia de columna folio en SAP
    sap_fac_cols = [c for c in df_sap.columns
                    if any(kw in c for kw in ['Factura', 'Folio', 'Doc.', 'Documento'])]
    sap_fac_col  = sap_fac_cols[0] if sap_fac_cols else 'Folio Factura'
    df_sap['FOLIO_JOIN'] = df_sap[sap_fac_col].apply(limpiar_pedido)

    df_final = df_sap.merge(df_bl[['PEDIDO_CLEAN', 'BL', 'FechaCreacion']], on='PEDIDO_CLEAN', how='left')

    if not df_analistas.empty and 'FOLIO_JOIN' in df_analistas.columns:
        df_final = df_final.merge(df_analistas[['FOLIO_JOIN', 'Creado por']], on='FOLIO_JOIN', how='left')
    else:
        df_final['Creado por'] = pd.NA

    if not df_analistas.empty and 'PEDIDO_JOIN' in df_analistas.columns:
        df_a2 = df_analistas[['PEDIDO_JOIN', 'Creado por']].rename(columns={'Creado por': 'Creado_por_Ped'})
        df_a2 = df_a2.drop_duplicates(subset=['PEDIDO_JOIN'])
        df_final = df_final.merge(df_a2, left_on='PEDIDO_CLEAN', right_on='PEDIDO_JOIN', how='left')
    else:
        df_final['Creado_por_Ped'] = pd.NA

    df_final = df_final.merge(df_asignacion[['CLIENTE_NORM', 'Responsable']], on='CLIENTE_NORM', how='left')
    df_final = df_final.merge(df_manual[['PEDIDO_CLEAN', 'MOTIVO']], on='PEDIDO_CLEAN', how='left')
    df_final = df_final.merge(
        df_dus[['PEDIDO_CLEAN', 'DUS_LEGALIZADO', 'FECHA_DUS', 'ESTADO_DUS_RAW']],
        on='PEDIDO_CLEAN', how='left'
    )
    df_final['DUS_LEGALIZADO'] = df_final['DUS_LEGALIZADO'].fillna(False)

    def calcular_responsable(row):
        if row['CLIENTE_NORM'] in terrestres_set:  return "🚚 Terrestre"
        if pd.notna(row.get('Creado por')):          return row['Creado por']
        if pd.notna(row.get('Creado_por_Ped')):      return row['Creado_por_Ped']
        if pd.notna(row.get('Responsable')):         return row['Responsable']
        return "Sin Asignar"

    df_final['RESPONSABLE_FINAL'] = df_final.apply(calcular_responsable, axis=1)

    # ── 6. KPIs y estatus ─────────────────────────────────
    print("[6/6] Calculando KPIs y generando output...")
    df_final = df_final.rename(columns={
        'NPedido':            'N° PEDIDO',
        'Nombre Solicitante': 'NOMBRE SOLICITANTE',
        'Fecha Factura':      'FECHA FACTURA',
        'FechaCreacion':      'FECHA CREACIÓN BL',
        'RESPONSABLE_FINAL':  'RESPONSABLE',
        'FECHA_DUS':          'FECHA DUS',
        'ESTADO_DUS_RAW':     'ESTADO DUS',
    })

    df_final['FECHA FACTURA']     = pd.to_datetime(df_final['FECHA FACTURA'],     errors='coerce')
    df_final['FECHA CREACIÓN BL'] = pd.to_datetime(df_final['FECHA CREACIÓN BL'], errors='coerce')

    # C-01: pasar FECHA_HOY como argumento a calcular_semaforo
    df_final['SEMÁFORO']      = df_final.apply(lambda r: calcular_semaforo(r, FECHA_HOY), axis=1)
    df_final['ESTATUS_FINAL'] = df_final.apply(determinar_estatus, axis=1)

    # DEMORA: días desde factura hasta hoy
    df_final['DEMORA'] = (FECHA_HOY - df_final['FECHA FACTURA']).dt.days
    df_final.loc[df_final['ESTATUS_FINAL'] == '📦 Pendiente Facturación', 'DEMORA'] = 0
    df_final.loc[df_final['DEMORA'] > 10000, 'DEMORA'] = 0

    # T. GESTIÓN: días factura → BL (eficiencia interna)
    df_final['T. GESTIÓN'] = (df_final['FECHA CREACIÓN BL'] - df_final['FECHA FACTURA']).dt.days

    # T. LEGALIZACIÓN: días factura → DUS legalizado (ciclo completo)
    df_final['T. LEGALIZACIÓN'] = (df_final['FECHA DUS'] - df_final['FECHA FACTURA']).dt.days
    df_final.loc[~df_final['DUS_LEGALIZADO'], 'T. LEGALIZACIÓN'] = pd.NA

    # Columnas de output en orden exacto
    columnas_output = [
        'ESTATUS_FINAL',
        'DEMORA',
        'T. GESTIÓN',
        'T. LEGALIZACIÓN',
        'N° PEDIDO',
        'NOMBRE SOLICITANTE',
        'RESPONSABLE',
        'FECHA FACTURA',
        'FECHA CREACIÓN BL',
        'FECHA DUS',
        'ESTADO DUS',
        'MOTIVO',
    ]
    columnas_output = [c for c in columnas_output if c in df_final.columns]
    df_export = df_final[columnas_output].copy()

    df_export.to_excel(OUTPUT_FILE, index=False)
    _aplicar_estilos(OUTPUT_FILE)

    total   = len(df_export)
    proc    = df_export['ESTATUS_FINAL'].str.contains('✅').sum()
    pend    = df_export['ESTATUS_FINAL'].str.contains('❌').sum()
    dus_leg = df_export['ESTATUS_FINAL'].str.contains('DUS Legalizado').sum()

    print(f"\n{'='*60}")
    print(f"  Total pedidos  : {total}")
    print(f"  Procesados     : {proc}")
    print(f"  DUS Legalizado : {dus_leg}")
    print(f"  Pendientes     : {pend}")
    print(f"  Output         : {OUTPUT_FILE}")
    print(f"{'='*60}")


if __name__ == "__main__":
    try:
        procesar_auditoria()
    except Exception as e:
        import traceback
        print(f"\n❌ ERROR CRÍTICO: {e}")
        traceback.print_exc()
